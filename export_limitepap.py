import openpyxl
import json
import re
import unicodedata

xlsx_file = 'Limite PAP Emendas 2025.xlsx'
json_file = 'limitespap.json'

def normalize(s):
    if not isinstance(s, str):
        return ''
    s = s.strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9_/ ]', '', s)
    return s

wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active

# Mostrar as 10 primeiras linhas para debug
print('Primeiras linhas da planilha:')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f'{i}: {row}')
    if i >= 10:
        break

# Procurar a linha do cabeçalho correto
header_row_idx = None
header = None
for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    row_norm = [normalize(cell) for cell in row]
    if 'UF' in row_norm and 'MUNICIPIO' in row_norm and 'IBGE' in row_norm and 'VALOR' in row_norm:
        header_row_idx = idx
        header = row
        break

if header_row_idx is None:
    print('Cabeçalho não encontrado!')
    exit(1)

header_norm = [normalize(h) for h in header]
print('Cabeçalho original:', header)
print('Cabeçalho normalizado:', header_norm)

col_names = {
    'UF': 'UF',
    'IBGE': 'IBGE',
    'MUNICIPIO': 'MUNICÍPIO',
    'VALOR': ' VALOR',
    'TIPO': 'TIPO',
}
col_map = {k: header_norm.index(normalize(v)) if normalize(v) in header_norm else None for k, v in col_names.items()}
print('Mapeamento de colunas:', col_map)

def parse_valor(valor):
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        valor = valor.replace('.', '').replace(',', '.')
        try:
            return float(valor)
        except Exception:
            return None
    return None

data = []
for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
    if not any(row):
        continue
    item = {
        'uf': row[col_map.get('UF')],
        'ibge': row[col_map.get('IBGE')],
        'municipio': row[col_map.get('MUNICIPIO')],
        'valor': parse_valor(row[col_map.get('VALOR')]),
        'tipo': row[col_map.get('TIPO')],
    }
    if item['tipo'] == 'PAP' and item['municipio']:
        data.append(item)

with open(json_file, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f'Dados exportados para {json_file} com sucesso!') 