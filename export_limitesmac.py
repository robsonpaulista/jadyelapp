import openpyxl
import json
import re
import unicodedata

# Nome dos arquivos
xlsx_file = 'Limite MAC Emendas 2025.xlsx'
json_file = 'limitesmac.json'

def normalize(s):
    if not isinstance(s, str):
        return ''
    s = s.strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9_/ ]', '', s)
    return s

# Carregar a planilha
wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active

# Procurar a linha do cabeçalho correto
header_row_idx = None
header = None
for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    row_norm = [normalize(cell) for cell in row]
    if 'UF' in row_norm and 'IBGE' in row_norm and 'ESTADO/MUNICIPIO' in row_norm and 'VALOR' in row_norm:
        header_row_idx = idx
        header = row
        break

if header_row_idx is None:
    print('Cabeçalho não encontrado!')
    exit(1)

header_norm = [normalize(h) for h in header]
print('Cabeçalho original:', header)
print('Cabeçalho normalizado:', header_norm)

# Nomes normalizados das colunas que queremos
col_names = {
    'UF': 'UF',
    'IBGE': 'IBGE',
    'ESTADO/MUNICIPIO': 'ESTADO/MUNICÍPIO',
    'VALOR': 'VALOR',
    'CNES': 'CNES',
    'NOME FANTASIA': 'NOME FANTASIA',
    'TIPO': 'TIPO',
}

col_map = {k: header_norm.index(normalize(v)) if normalize(v) in header_norm else None for k, v in col_names.items()}
print('Mapeamento de colunas:', col_map)

# Função para converter valor monetário para float
def parse_valor(valor):
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        valor = valor.replace('.', '').replace(',', '.')
        try:
            return float(valor)
        except Exception:
            return None
    return None

# Ler dados
data = []
for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
    if not any(row):
        continue
    item = {
        'uf': row[col_map.get('UF')],
        'ibge': row[col_map.get('IBGE')],
        'municipio': row[col_map.get('ESTADO/MUNICIPIO')],
        'valor': parse_valor(row[col_map.get('VALOR')]),
        'cnes': row[col_map.get('CNES')],
        'nome_fantasia': row[col_map.get('NOME FANTASIA')],
        'tipo': row[col_map.get('TIPO')],
    }
    # Só adiciona se for tipo MAC e município não vazio
    if item['tipo'] == 'MAC' and item['municipio']:
        data.append(item)

# Salvar em JSON
with open(json_file, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f'Dados exportados para {json_file} com sucesso!') 